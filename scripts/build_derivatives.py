"""Build CSV + XLSX derivatives + an index.html from the four JSON canonicals.

Canonical files (all live at the repo root):
    trust_urls.json          icb_urls.json
    trust-contacts.json      icb-contacts.json

For each canonical, generates three derivatives next to it:
    <stem>.csv               flattened CSV (arrays joined by ';', nulls -> empty string)
    <stem>.xlsx              same data as a single-sheet Excel file with frozen header

Plus a top-level:
    index.html               human-friendly landing page with links + entry counts + date

Run from anywhere; finds the repo root by file location.

Usage:
    py scripts/build_derivatives.py            # build into the same folder as the canonicals
    py scripts/build_derivatives.py <out_dir>  # build into a different folder
"""
from __future__ import annotations

import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. `pip install --user openpyxl`", file=sys.stderr)
    sys.exit(1)

REPO = Path(__file__).resolve().parent.parent
DATASETS = ["trust_urls", "icb_urls", "trust-contacts", "icb-contacts"]


def _flatten(entry: dict) -> dict:
    """Convert one JSON entry to a flat dict for CSV/XLSX.

    Arrays become semicolon-joined strings. Nulls become empty strings.
    For the URL files, the `names` array is split: first goes to `name`,
    rest to `alt_names`.
    """
    out = {}
    for k, v in entry.items():
        if k == "names":
            names = list(v or [])
            out["name"] = names[0] if names else ""
            out["alt_names"] = ";".join(names[1:])
            continue
        if isinstance(v, list):
            out[k] = ";".join(str(x) for x in v)
        elif v is None:
            out[k] = ""
        else:
            out[k] = v
    return out


def _column_union(rows: list[dict]) -> list[str]:
    """Build an ordered union of all keys seen across rows (preserves first-seen order)."""
    cols: list[str] = []
    for r in rows:
        for k in r:
            if k not in cols:
                cols.append(k)
    return cols


def _write_csv(rows: list[dict], cols: list[str], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, quoting=csv.QUOTE_MINIMAL, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def _write_xlsx(rows: list[dict], cols: list[str], path: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name 31-char limit

    # Header row
    bold = Font(bold=True)
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold

    # Data rows
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(cols, start=1):
            ws.cell(row=row_idx, column=col_idx, value=r.get(col_name, ""))

    # Freeze the header
    ws.freeze_panes = "A2"

    # Auto-size columns (capped at 60 chars to keep things readable)
    for col_idx, col_name in enumerate(cols, start=1):
        max_len = len(col_name)
        for r in rows:
            v = r.get(col_name, "")
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(path)


def build_one(stem: str, src_root: Path, out_root: Path) -> int:
    src = src_root / f"{stem}.json"
    data = json.loads(src.read_text(encoding="utf-8"))
    rows = [_flatten(e) for e in data]
    cols = _column_union(rows)

    out_root.mkdir(parents=True, exist_ok=True)
    _write_csv(rows, cols, out_root / f"{stem}.csv")
    _write_xlsx(rows, cols, out_root / f"{stem}.xlsx", sheet_name=stem)
    return len(data)


DATASET_LABELS = {
    "trust_urls":     "Trust board-papers URLs",
    "icb_urls":       "ICB board-papers URLs",
    "trust-contacts": "Trust press & FOI contacts",
    "icb-contacts":   "ICB press & FOI contacts",
}


def _human_size(path: Path) -> str:
    n = path.stat().st_size
    for unit in ("B", "KB", "MB"):
        if n < 1024:
            return f"{n:.0f} {unit}" if unit == "B" else f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} GB"


def write_index_html(counts: dict[str, int], out_root: Path) -> None:
    """Build a human-friendly landing page at the repo root.

    Lists the four datasets with download links for each format, current
    entry counts, file sizes, and last-built date. Static HTML, no JS.
    """
    today = datetime.now(timezone.utc).strftime("%-d %B %Y") if sys.platform != "win32" \
            else datetime.now(timezone.utc).strftime("%#d %B %Y")

    rows_html = []
    for stem, label in DATASET_LABELS.items():
        n = counts.get(stem, 0)
        sizes = {ext: _human_size(out_root / f"{stem}.{ext}") for ext in ("json", "csv", "xlsx")}
        rows_html.append(f"""
      <tr>
        <td><strong>{label}</strong></td>
        <td class="num">{n}</td>
        <td><a href="{stem}.json">JSON</a> <span class="size">({sizes['json']})</span></td>
        <td><a href="{stem}.csv">CSV</a> <span class="size">({sizes['csv']})</span></td>
        <td><a href="{stem}.xlsx">XLSX</a> <span class="size">({sizes['xlsx']})</span></td>
      </tr>""")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>NHS trust + ICB data</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            max-width: 820px; margin: 2rem auto; padding: 0 1rem; line-height: 1.55; color: #222; }}
    h1 {{ font-size: 1.6rem; margin-bottom: 0.2rem; }}
    .meta {{ color: #666; font-size: 0.9rem; margin-bottom: 1.5rem; }}
    table {{ border-collapse: collapse; width: 100%; margin: 1rem 0; }}
    th, td {{ text-align: left; padding: 0.5rem 0.7rem; border-bottom: 1px solid #eee; }}
    th {{ background: #f7f7f9; font-weight: 600; font-size: 0.9rem; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; color: #444; }}
    .size {{ color: #888; font-size: 0.85em; }}
    a {{ color: #c90c0f; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .note {{ background: #f7f7f9; padding: 0.7rem 1rem; border-left: 3px solid #c90c0f;
             margin: 1.5rem 0; font-size: 0.95rem; }}
    code {{ background: #f0f0f2; padding: 0.1rem 0.3rem; border-radius: 3px;
            font-family: ui-monospace, "Cascadia Code", Menlo, monospace; font-size: 0.9em; }}
  </style>
</head>
<body>
  <h1>NHS trust + ICB data</h1>
  <p class="meta">Public reference dataset for English NHS trusts and integrated care boards.<br>
  Last built: <strong>{today}</strong>. Refreshed automatically every Saturday.</p>

  <p>Four datasets, each available in three formats. <strong>XLSX</strong> is the most readable in
  Excel (with sortable columns); <strong>CSV</strong> is good for quick scripts; <strong>JSON</strong>
  is the canonical machine-readable format.</p>

  <table>
    <thead>
      <tr>
        <th>Dataset</th>
        <th>Entries</th>
        <th>Canonical</th>
        <th>Flat</th>
        <th>Spreadsheet</th>
      </tr>
    </thead>
    <tbody>{''.join(rows_html)}
    </tbody>
  </table>

  <div class="note">
    The data is aggregated from publicly-listed information on each organisation's website.
    Released under <a href="LICENSE">CC0 1.0</a> — use freely. The repository
    <a href="https://github.com/Davewest84/nhs-trust-icb-data">on GitHub</a> has commit
    history for every change; old URLs pointing at <code>nhs-board-papers-reader</code>
    (this repo's previous name) still redirect.
  </div>

  <h2 style="font-size: 1.15rem; margin-top: 2rem;">For tools and scripts</h2>
  <p>Direct raw URLs for the canonical JSON files:</p>
  <pre style="background: #f7f7f9; padding: 0.7rem 1rem; overflow-x: auto; font-size: 0.85rem;"><code>https://raw.githubusercontent.com/Davewest84/nhs-trust-icb-data/main/trust_urls.json
https://raw.githubusercontent.com/Davewest84/nhs-trust-icb-data/main/icb_urls.json
https://raw.githubusercontent.com/Davewest84/nhs-trust-icb-data/main/trust-contacts.json
https://raw.githubusercontent.com/Davewest84/nhs-trust-icb-data/main/icb-contacts.json</code></pre>

  <h2 style="font-size: 1.15rem; margin-top: 2rem;">Maintained by</h2>
  <p>Dave West at HSJ. Issues or corrections via
  <a href="https://github.com/Davewest84/nhs-trust-icb-data/issues">GitHub issues</a>
  or get in touch directly. The repository's <a href="README.md">README</a> has the full
  schemas, refresh cadence and maintenance details.</p>
</body>
</html>
"""
    (out_root / "index.html").write_text(html, encoding="utf-8")


def main() -> int:
    out_root = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else REPO
    print(f"Source canonicals: {REPO}")
    print(f"Output:           {out_root}")
    print()
    counts = {}
    for stem in DATASETS:
        n = build_one(stem, REPO, out_root)
        counts[stem] = n
        print(f"  {stem}: {n} entries  ->  {stem}.csv + {stem}.xlsx")
    write_index_html(counts, out_root)
    print(f"  index.html written")
    return 0


if __name__ == "__main__":
    sys.exit(main())
